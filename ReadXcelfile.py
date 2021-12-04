# Python program to read an excel file
 
# import openpyxl module
import openpyxl
 
# Give the location of the file
path2 = "/media/kornbotdev/8D60-165C/รายชื่อสำหรับ test ระบบ.xlsx"
path = "/media/kornbotdev/8D60-165C/demo.xlsx"

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>.
    # Getting the len of the data inside the array 
Row_len = []  
Column_len = []
Header_build = [] 
Index_tuber = []
dictcsv = {}
dict_complete_query = {} # Getting the dictionary of the data for comparation process 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active 
Tube_mem = []
for i in sheet_obj.iter_rows(max_row=0):

    #print(len(i))
    Row_len.append(len(i)) 
    #for r in range(0,len(i)):
    #   print(i[r],r)   
column=sheet_obj['A']
#print(column,len(column))
Column_len.append(len(column))
for memr in range(0,Column_len[0]):
            Tube_mem.append([])
#print(Tube_mem)
#print(Row_len[len(Row_len)-1],Column_len[0])

## Running the loop of the row and column function 
for k in range(1,Column_len[0]+1):
       for q in range(1,Row_len[len(Row_len)-1]+1):
                  #q = 2
                  tubeindex = sheet_obj.cell(row = k, column = q)
                  #print(k,q,tubeindex.value)
                  Tube_mem[k-1].append(tubeindex.value)
#print(Tube_mem)
row_header = Tube_mem[0]
for h in range(0,Column_len[0]+1):
         Header_build.append("P"+str(h))
#print("Header: ",Header_build)
for r in range(1,len(Tube_mem)):
         dictcsv[Header_build[r]] = Tube_mem[r]

#print(dictcsv)
for rew in list(dictcsv):
        print(rew,dictcsv.get(rew))
for wed in range(0,len(list(dictcsv))):
        
        Index_tuber.append(list(dictcsv.values())[wed][0])
for head in range(0,len(Index_tuber)):
        
        dict_complete_query[Index_tuber[head]]  = list(dictcsv.values())[head]

print("Patient data: ",dict_complete_query)
print("Tube index code: ",list(dict_complete_query))      
for i in range(0,len(list(dict_complete_query))):
              print(dict_complete_query.get(list(dict_complete_query)[i]))
