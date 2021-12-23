#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#Author:Chanapai Chuadchum
#Project:Covid detection GUI 
#release date:13/11/2021
import openpyxl # File converter 
from pyzbar import pyzbar 
import getpass 
import cv2  # Opencv for the camera qr code tracking and the covid 19 detection data 
import qrcode #Qr code generator with this code 
from printrun.printcore import printcore
from printrun import gcoder
from PyQt5.QtCore import Qt, QSize, QTimer, QThread
#import spidev # This library only working on the rpi hardware to control the gpio 
#import ws2812 # Getting the ws2813 communicate with the hardware direcly
from PyQt5 import QtCore, QtWidgets, uic,Qt,QtGui 
from PyQt5.QtWidgets import QApplication,QTreeView,QDirModel,QFileSystemModel,QVBoxLayout, QTreeWidget,QStyledItemDelegate, QTreeWidgetItem,QLabel,QGridLayout,QLineEdit,QDial,QComboBox,QPushButton 
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QPixmap,QIcon,QImage,QPalette,QBrush
#from pyqtgraph.Qt import QtCore, QtGui   #PyQt graph to control the model grphic loaded  
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
#import pyqtgraph.opengl as gl
import pandas as pd # Reading the pandas data from the csv data of the patient 
import csv #reading the csv file data 
import re
import os 
import sys
import serial # Serial protocol for connecting to the pheripheral devices like microcontroller 
import time
import getopt
import numpy
from numpy import sin, pi
import subprocess # subprocess for running the command inside the linux system 
config_Data = {"Top_catesian":{'x':200,'y':200},"Bottom_catesian":{'x':200,'y':200}} #Configuretion data in json for the stepper motor control step co-ordination
#Setting the current position in array 
#Top cam catsian robot 
xt_array = [] 
yt_array =[]
#Bottom cam catesian robot 
xb_array = [] 
yb_array = []
Devices_mem = [] # Devices mem the serial device
#Calibration interval from the auto calibration 
Covid_interval = []
#referent covid status
Covid_status = [] #current covid19 status on the testtube detected on each position and colllect into the json data created on the list 
seriallist = os.listdir("/sys/class/tty") 
serialmem1 = []

Patientpersonaldata = ['index no.','patient name','age'] #Getting the patients personal data as the key for the system 
Patientdata = {} #Getting the json file generated for the qrcode generator 
name_patient =[]  #Getting the name of the patient 
age_patient = []  #Getting the age of the patient 
username = getpass.getuser() #Getting the host of the user 
PATHDIR = "/media/" + username #Getting the path from the list of this data 
path_for_maketubeindex = "/home/"+username+"/"+"tubeindex/" #Path for making the tube index for the directory
tubeindex_mem_delete = [] # Deleted path for the file of tubeindex 
path_for_patient = "/home/"+username+"/"+"patientfiles/" #Path for making the patient directory 
patient_mem_delete = [] # Delete path fo the file of the patients 
listPATH_drive = os.listdir(PATHDIR) #Getting the data inside the list of the path

dirmem_ext = [] #Getting the external drive list data  
external_file = [] #Getting the external file in the directory
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 # Camera devices get index 
cameradevmem = [] #getting the camera mem data inside the list of the function 
cam_index = [] #Getting the camera device index data 
Getcam = []
Index_cam =[] #Getting the index camera mem inside the list
cam_num = [] # Getting the camera number 
Qr_listdata = [] #Getting the qr code list data function
Qr_ref_motion = [] #Getting the qrcode position data for reference motion tracking 
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# Qr code data generator  
Dictheader = {}
Qr_prepdata = {}
Qr_readydata = {}
remove_patient_nan = []
patientdata_edit = [] 
array_patient_info = [[],[],[],[],[],[]]
array_patient_edit = [[],[],[],[],[],[]]
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# Path data reference 
PATH_External = []
PATH_Internal = [] 
Patient_PATH = [] # Save the patient path 
Tube_index_path = [] #Save the tube index path
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
   # Reference array for the data of the precise detection mod 
Mem_process_Qr = [] #Getting the detected data from the camera qr code detection to avoiding the same position detection
Mem_covid_dec = [] #Getting the covid data detection avoid same position detection 
    # Getting the len of the data inside the array 
Row_len = []  
Column_len = []
Header_build = [] 
Index_tuber = []
dictcsv = {}
dict_complete_query = {} # Getting the dictionary of the data for comparation process
Tube_mem = []
   # Create the directory for the patient and tube index
Row_patients_len = []  
Column_patients_len = []
Header_patients = []
Index_patients = []
dictpatientcsv = {}
dict_patients_query = {} # Getting the data of the patient query from the tube index code 
Patients_mem = []
Groupping_Patient = [[],[],[],[],[],[],[],[]] # Grouping the patient data 
Ref_group = [] # Reference group for the data of the patient 
Patient_qr_member = {} # Getting the member of the patient on the qr code detection 
patient_mem_list = []
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
   #Getting the data from the covid detector status 
Refferrence_len_array = [] # Ref array len to store the status of
Covid_ref_status = [] # Getting the covid status 
Covid_tubeindex_liststatus = {} # Getting the tube index status containing x,y position of the covid detection  

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
try: 
   print("Creating the tubeindex and patient directory")
   mode = 0o775    # Mode for making the chmod permission 775   
   os.mkdir(path_for_maketubeindex,mode) #Tube index path 
   os.mkdir(path_for_patient,mode) #Patient path
   #List the path file for the data of the tube and patient matching function  
   tubeindex_list = os.listdir(path_for_maketubeindex)
   patient_list = os.listdir(path_for_patient) 
except:
    print("Directory was created")
tubeindex_list = os.listdir(path_for_maketubeindex)
patient_list = os.listdir(path_for_patient)
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# Camera , SerialUSB , pheripheral device detect must running in the multitasking programming to automaticly detect the devices 
# List the camera webcam available connected with the computer 
try:
  listcamera = subprocess.check_output("v4l2-ctl --list-devices",shell=True)
  print("Camera devices listing")
  print(listcamera.decode()) #decode the list the camera device
  camlist = listcamera.decode().split("/dev/")
  for re in range(0,len(camlist)):
          print(camlist[re])
          cameradevmem.append(camlist[re])
  print(cameradevmem,len(cameradevmem))
  for indexcam in range(0,len(cameradevmem)-1):
              cam_index.append(cameradevmem[indexcam]) #Getting the index unsorted value 
  print(cam_index)
  #Getting the video camera with modulo technique 
  for qs in range(0,len(cam_index)):
            modulo = qs%2  #Getting the modulo calculate inside the data 
            if modulo == 1:
                Getcam.append(cam_index[qs]) #Getting the camera devices 
  print(Getcam[0])
  for w in range(0,len(Getcam)):
         print(str(Getcam[w])) #Getting the camera video list devices 
         Index_cam.append(str(Getcam[w]))
  print("Index camera: ",Index_cam)
  for wr in range(0,len(Index_cam)):
          print(list(Index_cam[wr]))
          listcamdat = list(Index_cam[wr])
          print(listcamdat[len(listcamdat)-3]) 
          indexcamdat = listcamdat[len(listcamdat)-3]
          cam_num.append(indexcamdat) 
  print("Camera list available ID: ",cam_num)
except: 
    print("No such camera device found!")

print(Index_cam) #Show the index cam list to using inside the terminal 

class MainWindow(QtWidgets.QMainWindow):

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        
        #Load the UI Page
        uic.loadUi('coviddetector.ui', self)
        self.setWindowTitle('Covid detector')
        p = self.palette()
        p.setColor(self.backgroundRole(), QtCore.Qt.white)
        self.setPalette(p)
        #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
        # Covid detector page 
        #label pixmap  camera 
        self.camera = self.findChild(QLabel,"label_16")
        self.camera.setFixedSize(531,441)
        #Button function front 
        self.pushButton_6.clicked.connect(self.Rundetect) #Running the detector mode 
        self.pushButton_5.clicked.connect(self.Reportresult) #Report the result into the csv and pdf file 
        self.pushButton_9.clicked.connect(self.Qrgenerator_fromlist) #Getting the list of the customer to add in the list and concatinate the function of the qr code 
        self.pushButton_10.clicked.connect(self.Visual1)
        self.pushButton_3.clicked.connect(self.Clear_Selected_file) #Clear the selected file 
        self.pushButton_4.clicked.connect(self.Clear_all_files) #Clear all the file selected from the patient and tube index inside the list 

        #Combobox external drive data 
        self.combo_external = self.findChild(QComboBox,"comboBox_2") #Getting the combobox for the external for write and read file storage 
        self.combo_external.addItem("No drive") #List of status no drive detected 
        #List path data drive 
        for r in range(0,len(listPATH_drive)):
                 print(listPATH_drive)
                 for q in range(0,len(listPATH_drive)):
                             dirmem_ext.append(PATHDIR+"/"+listPATH_drive[q]) #Getting the drive file
        self.combo_external.addItems(dirmem_ext) #Getting the list of the external drive
        if listPATH_drive !=[]:
            try:
                #for w in range(0,len(dirmem_ext)):
                #       fileinside = os.listdir(dirmem_ext[w])
                #       print(fileinside) 
                #       for qw in range(0,len(fileinside)): #Getting the file inside the listdir 
                fileinside = os.listdir(PATHDIR+"/"+listPATH_drive[0]) #Get the file inside the directory 
                for qw in range(0,len(fileinside)):        
                        external_file.append(fileinside[qw]) #Getting the file append inside the external list 
                        
            except: 
                print("No drive found")
        #Setting the file from the drive inside directory 
        self.combo_external.activated[str].connect(self.Externaldata)      
        #List file inside the selected drive 
        self.combo_internal = self.findChild(QComboBox,"comboBox_3") #Getting the file inside the external drive selected 
        self.combo_internal.addItem("No-file") #No file in the case not selected 
        self.combo_internal.addItems(external_file) #Getting the list file extrackted from the list 
        self.combo_internal.activated[str].connect(self.Internalfile) #Getting the internal file 
        #Combobox serial communication with the hardware catesian robot 
        self.comboserial = self.findChild(QComboBox,"comboBox") #Get the serial communication from the combobox 
        self.comboserial.addItem("Non-serial") # Getting the item list on the serial search function 
        #Setting the combobox for the tube index path 
        self.combotube = self.findChild(QComboBox,"comboBox_4") #Getting the path file for the index 
        self.combotube.addItem("No-file")
        self.combotube.addItems(external_file)
        self.combotube.activated[str].connect(self.Tubeindex_path) #Sending file name to assembly with the path 
       

        self.comboclearpatient = self.findChild(QComboBox,"comboBox_6") # Select the file in the created directory to clear the data 
        self.comboclearpatient.addItem("No-file")
        self.comboclearpatient.addItems(patient_list) #Getting the data from the patient list 
        self.comboclearpatient.activated[str].connect(self.Patient_clear_file) # combobox data visualize on the list 
        
        for i in range(0,len(seriallist)):
               
              if len(str(seriallist[i]).split("USB")) >= 2:

                        serialmem1.append(str(seriallist[i])) 
                        
              if len(str(seriallist[i]).split("ACM")) >= 2: 

                        serialmem1.append(str(seriallist[i])) 

        self.comboserial.addItems(serialmem1) #Get the serial data       
        self.comboserial.activated[str].connect(self.Serialfunc)      
        
        #Label for the function of the the video frame display on the opencv covid 19 detection function 
        
        #Top camera #QR code detection 
        self.camera2 = self.findChild(QLabel,"label_8")
        self.camera2.setFixedSize(541,431)

        self.camera3 = self.findChild(QLabel,"label") 
        self.camera3.setFixedSize(541,431) 

        #Button control 
        self.pushButton_7.clicked.connect(self.Homeposition_top) #Getting the home position for the top catesian robot 
        self.pushButton.clicked.connect(self.Y_axis_neg_top)     #Getting the Y axis negative control 
        self.pushButton_14.clicked.connect(self.Y_axis_pos_top) #Getting the Y axis positive control 
        self.pushButton_15.clicked.connect(self.X_axis_pos_top) #Getting the X axis positive control
        self.pushButton_16.clicked.connect(self.X_axis_neg_top) #Getting the X axis negative control 
        
        self.camon2 = self.findChild(QPushButton,"pushButton_11") 
        self.camon2.clicked.connect(self.Visual2) #Getting the camera top view on detecting the QRcode and calibrate position 
        #Dial control 
        self.dial1 = self.findChild(QDial,"dial") # Getting the dial 1 for the function of the light intensity control 
        self.dial1.setMinimum(0)
        self.dial1.setMaximum(255) # PWM output control function
        self.dial1.setValue(0)
        self.dial1.valueChanged.connect(self.Light_intense_top) #Top light intensity control 
        self.dial2 = self.findChild(QDial,"dial_2") #Getting the dial 2 for the function of the wavelength control function 
        #X-axis control herizontal slider  
        self.slider_xt = self.findChild(QSlider,"horizontalSlider") #Getting the X axis function of the slider 
        self.slider_xt.setMinimum(0)
        self.slider_xt.setMaximum(config_Data.get('Top_catesian').get('x')) #Getting the maximum value of the X config data on the catesian robot        
        self.slider_xt.valueChanged.connect(self.Top_xt)
        #Y-axis control vertical slider 
        self.slider_yt = self.findChild(QSlider,"verticalSlider") #Getting the Y axis function of the slider to working 
        self.slider_yt.setMinimum(0) 
        self.slider_yt.setMaximum(config_Data.get('Top_catesian').get('y')) #Getting the maximum value of the Y config data on the catesian robot 
        self.slider_yt.valueChanged.connect(self.Top_yt)
        
        # Setting the slider function controller
        #Bottom camera 
        #Button control 
        self.pushButton_8.clicked.connect(self.Homeposition_bottom) 
        self.pushButton_2.clicked.connect(self.Y_axis_neg_bottom)  # Getting the Y negative axis control 
        self.pushButton_19.clicked.connect(self.Y_axis_pos_bottom) # Getting the Y positive axis control 
        self.pushButton_17.clicked.connect(self.X_axis_pos_bottom) # Getting the X positive axis control 
        self.pushButton_18.clicked.connect(self.X_axis_neg_bottom) # Getting the X negative axis control 

        self.dial3 = self.findChild(QDial,"dial_3") #Getting the dial 3 for the function of light intensity control 
        self.dial3.setMinimum(0)
        self.dial3.setMaximum(255) # PWM output control function
        self.dial3.setValue(0)
        self.dial3.valueChanged.connect(self.Light_intense_bottom) #Top light intensity control
        self.dial4 = self.findChild(QDial,"dial_4") #Getting the dial 4 for the function of the wavelength control function 
        self.slider_xb = self.findChild(QSlider,"horizontalSlider_2") 
        self.slider_xb.setMinimum(0) 
        self.slider_xb.setMaximum(config_Data.get('Bottom_catesian').get('x')) #Getting the config of the x axis at catesian robot 
        self.slider_xb.valueChanged.connect(self.Bottom_xb)

        self.slider_yb = self.findChild(QSlider,"verticalSlider_2")  #Getting the config of the y axis at catsian robot 
        self.slider_yb.setMinimum(0)
        self.slider_yb.setMaximum(config_Data.get('Bottom_catesian').get('y'))
        self.slider_yb.valueChanged.connect(self.Bottom_yb)
        #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
        #Pause and Resume button 
        self.pushButton_12.clicked.connect(self.Pausemotion) # Setting the pause motion control 
        self.pushButton_13.clicked.connect(self.Resumemotion) # Setting the resume motion control 


    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    #Camera 
    def Visual1(self):
            #Camerafunc(self.camera,0,451,441)
             self.Worker1 = Worker1()
             self.Worker1.start()
             self.Worker1.ImageUpdate.connect(self.ImageUpdateSlot)
    def Visual2(self):
            #Camerafunc(self.camera,0,451,441)
             self.Worker2 = Worker2()
             self.Worker2.start()
             self.Worker2.ImageUpdate.connect(self.ImageUpdateSlot2)
    
    # Combobox function 
    def Patient_clear_file(self,text):
        print("Clear the patients file name: ",text) # Getting the file name to make the path for clear the file 
        for r in range(0,1):
             self.comboclearpatient.clear()
             self.comboclearpatient.addItem("No-file")
             self.comboclearpatient.addItems(patient_list) 
             if patient_mem_delete != []:
                 try:
                    if len(patient_mem_delete) > 1:
                            
                             patient_mem_delete.remove(patient_mem_delete[0])
                             listtemsPatient = self.comboclearpatient.allItems() 
                             print(listtemsPatient)
                 except:
                     print("File in directory was deleted")
             if patient_mem_delete == []: 
                       
                          patient_mem_delete.append(path_for_patient+text)
   
    # Button Function of the covid detect page
    def Clear_Selected_file(self):
        print("Delete selected file")
        for r in range(0,1):
             if os.path.exists(tubeindex_mem_delete[0]):
                   
                   # removing the file using the os.remove() method
                   try:
                       os.remove(tubeindex_mem_delete[0])
                       self.comboclearpatient.clear()
                       self.comboclearpatient.addItem("No-file")
                   except:
                       print("File not found in the directory")
             if os.path.exists(patient_mem_delete[0]):
                   # removing the file using the os.remove() method
                   try:
                       os.remove(patient_mem_delete[0])
                       self.comboclearpatient.clear()
                       self.comboclearpatient.addItem("No-file")
                   except:
                       print("File not found in the directory")
    def Clear_all_files(self):
        print("Delete all files")
        for r in range(0,1):
               
             if os.path.exists(patient_mem_delete[0]):
                 
                # removing the file using the os.remove() method
                try:    
                   os.remove(patient_mem_delete[0])
                except: 
                   print("File not found in the directory")
    def Rundetect(self):
              print("Run detection") # Running the covid detection camera here to activate recoding the function with motion control catesian robot
              try:
                     pnp_serial = printcore('/dev/'+str(Devices_mem[0]),115200) #fixed baudrate  using the serial communication to control the gcode core motion 
                     while not pnp_serial.online:
                               time.sleep(0.1)  
                     pnp_serial.send_now("G0 X0 Y0") # Home position 
                     pnp_serial.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
                     pnp_serial.send_now("M302 S0")
                     pnp_serial.send_now("M106 S190")
                     pnp_serial.send_now("G0 X100 Y200")
                     pnp_serial.send_now("G0 E050")
                     pnp_serial.send_now("G0 Z170")
                     pnp_serial.send_now("G0 Z0")
                     pnp_serial.send_now("G0 E00")
                     pnp_serial.send_now("G0 E1180")
                     pnp_serial.send_now("G0 X0 Y0")
                     pnp_serial.send_now("M106 S0")
                     pnp_serial.pause() # use these to pause/resume the current print
                     pnp_serial.resume() 
                     pnp_serial.disconnect() # this is how you disconnect from the printer once you are done. This will also stop running prints.

              except:
                 print("Error serial connection loss on hardware")
    def Reportresult(self):
          print("Report result") # Report pdf file from the csv data frame 

    def Qrgenerator_fromlist(self):
          print("Qrcode generator")
          #Running the function of the patient data in the loop 
          #img = qrcode.make('{'+'name'+":"+"kornbot"+'}')  #Adding the data list from the add list function here
          #type(img)  # qrcode.image.pil.PilImage
          #img.save("patientname.png") #Getting the path from the text input in the combobox 
          #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
          #Getting the data from pandas library and running the data in here
          try:
             print("Path patient file: ",PATH_External[0]+"/"+PATH_Internal[0]) # Generate the main path of the storage devices and file for extraction 
             internalfile = PATH_Internal[0]
             nameof_internal = internalfile.split(".")[0]
             extention_internal = internalfile.split(".")[1] 

          except:
              print("No file input from combobox")  
          try:
             print("Path tube index file: ",PATH_External[0]+"/"+Tube_index_path[0]) #Generate the main path 
             tubeindexfile = Tube_index_path[0]
             nameof_tubeindex = tubeindexfile.split(".")[0]
             extention_tubeindex = tubeindexfile.split(".")[1]

          except:
              print("No file input from combobox")
          #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
          # Detecting the file type and convert file including matching the id index with the list of the patient id 
          """
          internalfile = PATH_Internal[0]
          tubeindexfile = Tube_index_path[0]
          #Check if file type is csv and xlsx if xlsx then convert into csv file 
          nameof_internal = internalfile.split(".")[0]
          extention_internal = internalfile.split(".")[1] 
          nameof_tubeindex = tubeindexfile.split(".")[0]
          extention_tubeindex = tubeindexfile.split(".")[1]
          """
          try:
             if extention_internal == 'xlsx':
               try:
                    print("Start convert file extention to xlsx")
                    Path_dataframe_internal = PATH_External[0]+"/"+PATH_Internal[0] 
                    print(Path_dataframe_internal)
                    wb_obj = openpyxl.load_workbook(Path_dataframe_internal)
                    sheet_obj = wb_obj.active 
                    for i in sheet_obj.iter_rows(max_row=0):
                                 #print(len(i))
                                 Row_patients_len.append(len(i))    
                    column=sheet_obj['A']
                    #print(column,len(column))
                    Column_patients_len.append(len(column))
                    for memr in range(0,Column_patients_len[0]):
                               Patients_mem.append([])
                    #print(Tube_mem)
                    #print(Row_len[len(Row_len)-1],Column_len[0])
                    ## Running the loop of the row and column function 
                    for k in range(1,Column_patients_len[0]+1):
                          for q in range(1,Row_patients_len[len(Row_patients_len)-1]+1):
                             #q = 2
                             tubeindex = sheet_obj.cell(row = k, column = q)
                             #print(k,q,tubeindex.value)
                             Patients_mem[k-1].append(tubeindex.value)
                    #print(Tube_mem)
                    row_header = Patients_mem[0]
                    for h in range(0,Column_patients_len[0]+1):
                           Header_patients.append("P"+str(h))
                    #print("Header: ",Header_build)
                    for r in range(1,len(Patients_mem)):
                            dictpatientcsv[Header_patients[r]] = Patients_mem[r]
                    #print(dictcsv)
                    for rew in list(dictpatientcsv):
                           print("Patients data: ",rew,dictpatientcsv.get(rew))
                    for wed in range(0,len(list(dictpatientcsv))):
                           Index_patients.append(list(dictpatientcsv.values())[wed][0])
                    for head in range(0,len(Index_patients)):
                              dict_patients_query[Index_patients[head]]  = list(dictpatientcsv.values())[head]

                    print("Patient list data: ",dict_patients_query)
                    print(Patients_mem)
                             
                                   
               except: 
                   print("No file input from combo box")   
          except:
              print("No file input from combobox")
          try:
            if extention_tubeindex == 'xlsx':
                try:
                    print("Start convert file extention to xlsx") 
                    Path_dataframe_tubeindex = PATH_External[0]+"/"+Tube_index_path[0]
                    print(Path_dataframe_tubeindex)
                    wb_obj = openpyxl.load_workbook(Path_dataframe_tubeindex)
                    sheet_obj = wb_obj.active 
                    for i in sheet_obj.iter_rows(max_row=0):
                                 #print(len(i))
                                 Row_len.append(len(i))    
                    column=sheet_obj['A']
                    #print(column,len(column))
                    Column_len.append(len(column))
                    for memr in range(0,Column_len[0]):
                               Tube_mem.append([])
                    #print(Tube_mem)
                    #print(Row_len[len(Row_len)-1],Column_len[0])
                    ## Running the loop of the row and column function 
                    for k in range(1,Column_len[0]+1):
                          for q in range(1,Row_len[len(Row_len)-1]+1):
                             #q = 2
                             tubeindex = sheet_obj.cell(row = k, column = q)
                             #print(k,q,tubeindex.value)
                             Tube_mem[k-1].append(tubeindex.value)
                    #print(Tube_mem)
                    row_header = Tube_mem[0]
                    for h in range(0,Column_len[0]+1):
                           Header_build.append("P"+str(h))
                    #print("Header: ",Header_build)
                    for r in range(1,len(Tube_mem)):
                            dictcsv[Header_build[r]] = Tube_mem[r]
                    #print(dictcsv)
                    for rew in list(dictcsv):
                           print(rew,dictcsv.get(rew))
                    for wed in range(0,len(list(dictcsv))):
                           Index_tuber.append(list(dictcsv.values())[wed][0])
                    try:
                        for head in range(0,len(Index_tuber)):
                              print(list(dictcsv.values())[head],head)
                              dict_complete_query[Index_tuber[head]]  = list(dictcsv.values())[head]
                              
                        print(dict_complete_query)

                    except: 
                         print("Out of range")
                except: 
                   print("No file input from combo box")
          except:
              print("No file input")      
          else: 
             print("File extension not match any cases please using support file like .csv and .xlsx")          
    #Button Function of the Top camera 
    def Homeposition_bottom(self):
           print("Home position bottom")
           try:
              print(Devices_mem[0])
              if Devices_mem[0] != "Non-serial": 
                p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
                #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
                #gcode = gcoder.LightGCode(gcode)
                # startprint silently exits if not connected yet
                while not p.online:
                     time.sleep(0.1)
                p.send_now("G0 Z0") # Setting the home position 
                p.send_now("G0 E0") # Setting the home position 
                p.resume() # Pause the stepper motor
                p.disconnect() 
           except:
                 print("Serial device not found") 
  
          
    #Dial data 
    def Light_intense_top(self): 
          print("Light intensity top: ",self.dial1.value())
          self.lcd1 = self.findChild(QLCDNumber,"lcdNumber")
          self.lcd1.display(self.dial1.value())
          self.lcd1.setStyleSheet("""QLCDNumber { background-color: black; }""")
          
    def Light_intense_bottom(self):
          print("Light intensity bottom: ",self.dial3.value())
          self.lcd2 = self.findChild(QLCDNumber,"lcdNumber_5") 
          self.lcd2.display(self.dial3.value())
          self.lcd2.setStyleSheet("""QLCDNumber { background-color: black; }""")
          #spi = spidev.SpiDev()
          #spi.open(0,0)  
          mode = 'loop' # Getting the mode setting function
          #WS2812_controller(spi,mode,nLED,intensity) # Getting the light intensity input function to control the ws2812 rgb control 
    #Button Function of the Bottom camera 
    def Homeposition_top(self):
          print("Home position top")
          try:
              print(Devices_mem[0])
              if Devices_mem[0] != "Non-serial": 
                p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
                #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
                #gcode = gcoder.LightGCode(gcode)
                # startprint silently exits if not connected yet
                while not p.online:
                     time.sleep(0.1)
                p.send_now("G0 Z0 E00")    
                p.resume() # Pause the stepper motor
                p.disconnect() 
          except:
                 print("Serial device not found") 
    def Y_axis_neg_top(self):
        print("-Y: ",-yt_array[0]) # Array control the y neg axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"E-0"+str(yt_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
        p.send_now("M302 S0")
        #p.send_now("M106 S190")
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()
        
    def Y_axis_pos_top(self):
        print("+Y: ",yt_array[0])  # Array control the y pos axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"E0"+str(yt_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
        p.send_now("M302 S0")
        #p.send_now("M106 S190")
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()
    def X_axis_pos_top(self):     
        print("+X: ",xt_array[0])  # Array control the x pos axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"Z"+str(xt_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect() 

    def X_axis_neg_top(self):
        print("-X: ",-xt_array[0]) # Array control the x neg axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"Z-"+str(xt_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    # Bottom button controller 
    def Y_axis_neg_bottom(self): 
        print("-Y: ",-yb_array[0]) # Array control the y neg axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"Y-"+str(yb_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
        p.send_now("M302 S0")
        #p.send_now("M106 S190")
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()
    def Y_axis_pos_bottom(self):
        print("Y: ",yb_array[0]) # Array control the y neg axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"Y"+str(yb_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
        p.send_now("M302 S0")
        #p.send_now("M106 S190")
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()
    def X_axis_pos_bottom(self):
        print("X: ",-xb_array[0]) # Array control the y neg axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"X"+str(yb_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
        p.send_now("M302 S0")
        #p.send_now("M106 S190")
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()
    def X_axis_neg_bottom(self):
        print("-X: ",-xb_array[0]) # Array control the y neg axis 
        p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
        #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
        #gcode = gcoder.LightGCode(gcode)
        # startprint silently exits if not connected yet
        gcode_send_Xt = "G0 "+"X-"+str(xb_array[0])
        while not p.online:
                time.sleep(0.1)
        p.send_now("M302 P0") # this will send M105 immediately, ahead of the rest of the print
        p.send_now("M302 S0")
        #p.send_now("M106 S190")
        p.send_now(gcode_send_Xt)
        p.pause()
        p.resume()
        p.disconnect()

    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
               #Running the function of the value 
    def Top_xt(self):
        print("top_xt: ",self.slider_xt.value())  #Getting the value on each axis to save the value into the an array of each axis 
        if len(xt_array) >1:
             xt_array.clear() # Clear out the array if the array is more than one 
        xt_array.append(self.slider_xt.value())
        print(xt_array,len(xt_array))
        self.lcd3 = self.findChild(QLCDNumber,"lcdNumber_3")
        self.lcd3.display(self.slider_xt.value()) #Getting the top slider of X axis
        self.lcd3.setStyleSheet("""QLCDNumber { background-color: black; }""")
        Gcode_Xt_motion(xt_array)
    def Top_yt(self):
        print("top_yt: ",self.slider_yt.value())   
        if len(yt_array) >1:
            yt_array.clear() #Clear out the array 
        yt_array.append(self.slider_yt.value())
        print(yt_array,len(yt_array))
        self.lcd4 = self.findChild(QLCDNumber,"lcdNumber_4")
        self.lcd4.display(self.slider_yt.value()) #Getting the top slider of Y axis
        self.lcd4.setStyleSheet("""QLCDNumber { background-color: black; }""")
        Gcode_Yt_motion(yt_array) 

    def Bottom_xb(self):
        print("bottom_xb: ",self.slider_xb.value())   
        if len(xb_array) >1:
             xb_array.clear() #clear out the array 
        xb_array.append(self.slider_xb.value())
        print(xb_array,len(xb_array))
        self.lcd7 = self.findChild(QLCDNumber,"lcdNumber_7")
        self.lcd7.display(self.slider_xb.value()) #Getting bottom slider of X axis
        self.lcd7.setStyleSheet("""QLCDNumber { background-color: black; }""")
        Gcode_Xb_motion(xb_array)
    def Bottom_yb(self):
        print("bottom_yb: ",self.slider_yb.value())   
        if len(yb_array) >1:
            yb_array.clear() #clear out the array 
        yb_array.append(self.slider_yb.value()) 
        print(yb_array,len(yb_array))
        self.lcd8 = self.findChild(QLCDNumber,"lcdNumber_8") 
        self.lcd8.display(self.slider_yb.value()) #Getting bottom slider of Y axis 
        self.lcd8.setStyleSheet("""QLCDNumber { background-color: black; }""")
        Gcode_Yb_motion(yb_array)
    def Serialfunc(self,text,): #Getting the text from the list file of the serial
        for i in range(0,1): 
              for i in range(0,len(seriallist)):
               
                  if len(str(seriallist[i]).split("USB")) >= 2:

                            serialmem1.append(str(seriallist[i])) 
                        
                  if len(str(seriallist[i]).split("ACM")) >= 2: 

                            serialmem1.append(str(seriallist[i])) 
              print("serial_selected",text)
                
              if Devices_mem == []:
                    print("Mem serial port",str(text)) 
                    Devices_mem.append(text)
              if len(Devices_mem) > 1:
                       Devices_mem.remove(Devices_mem [0])  #remove the first one out from the list to get only the latest list 
              try:
                if text != "Non-serial":
                     pnp_serial = printcore('/dev/'+str(text),115200) #fixed baudrate  using the serial communication to control the gcode core motion 
                     while not pnp_serial.online:
                               time.sleep(0.1)  
                #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
                #Do process processing here for the G-code command running the fully autonomouse code on the robot 
                
              except:
                  print("Serial connection fail please recheck the serial communication at the pnp machine") 
    def Externaldata(self,text):
        for i in range(0,1):
              try:
                 print("External data: ",text) #the text is the actual directory of the drive 
                 try:
                    if PATH_External !=[]:
                       if PATH_External > 1:
                            PATH_External.remove(PATH_External[0])
                    if PATH_External == []:
                         PATH_External.append(text)
                 except: 
                     print("External error non detected")
              except:
                 print("No drive found") 
    def Internalfile(self,text):
        for i in range(0,1):
              try:
                  print("Internal data: ",text)
                  try: 
                     if PATH_Internal !=[]:
                       if PATH_Internal > 1:
                            PATH_Internal.remove(PATH_Internal[0])
                     if PATH_Internal == []:
                         PATH_Internal.append(text)
                  except: 
                       print("Internal error non detected")
              except:
                  print("No file found!") 
    def Tubeindex_path(self,text):
        for i in range(0,1):
              self.combotube.addItems(external_file)
              try:
                  print("Tube index path: ",text)  
                  try: 
                     if Tube_index_path !=[]:
                       if Tube_index_path > 1:
                            Tube_index_path.remove(Tube_index_path[0])
                     if Tube_index_path == []:
                         Tube_index_path.append(text)
                  except: 
                       print("Internal error non detected") 
              except:
                 print("Path not found external storage")
    def ImageUpdateSlot(self, Image):
            self.pixmap = QPixmap.fromImage(Image)
            self.camera.setPixmap(self.pixmap) 
            self.camera2.setPixmap(self.pixmap)        
    def ImageUpdateSlot2(self, Image):
            self.pixmap = QPixmap.fromImage(Image)
            self.camera3.setPixmap(self.pixmap)
    def Pausemotion(self):
             print("Pause stepper motor")
             try:
              print(Devices_mem[0])
              if Devices_mem[0] != "Non-serial": 
                p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
                #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
                #gcode = gcoder.LightGCode(gcode)
                # startprint silently exits if not connected yet
                while not p.online:
                     time.sleep(0.1)
                p.pause() # Pause the stepper motor
                p.resume()
                p.disconnect() 
             except:
                 print("Serial device not found")
    def Resumemotion(self):
             print("Resume stepper motor")
             try:
              print(Devices_mem[0])
              if Devices_mem[0] != "Non-serial": 
                p=printcore('/dev/'+str(Devices_mem[0]), 115200) # or p.printcore('COM3',115200) on Windows
                #gcode=[i.strip() for i in open('filename.gcode')] # or pass in your own array of gcode lines instead of reading from a file
                #gcode = gcoder.LightGCode(gcode)
                # startprint silently exits if not connected yet
                while not p.online:
                     time.sleep(0.1)
                p.resume() # Pause the stepper motor
                p.disconnect() 
             except:
                 print("Serial device not found") 
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#Motion control function     

def Gcode_Xt_motion(Xt): #Getting the motion control function for the catesian robot 
        print(Xt) #Getting the value from each axis input from each catesian robot 
        print("G0 "+"X"+str(Xt[0]))
        
def Gcode_Yt_motion(Yt): #Getting the motion control function for the catesian robot 
        print(Yt) #Getting the value from each axis input from each catesian robot 
        print("G0 "+"E0"+str(Yt[0]))
        
def Gcode_Xb_motion(Xb): #Getting the motion control function for the catesian robot 
        print(Xb) #Getting the value from each axis input from each catesian robot 
        print("G0 "+"X"+str(Xb[0]))
        
def Gcode_Yb_motion(Yb): #Getting the motion control function for the catesian robot 
        print(Yb) #Getting the value from each axis input from each catesian robot 
        print("G0 "+"Y"+str(Yb[0]))
        
def WS2812_controller(spi,mode,nLED,intensity): # Getting the data int the function to control the light intensity and mode 
      mode_list = ['loop','wave','npimage'] # Getting the mode list to control the function here 
      if mode in mode_list: 
            if mode == 'loop':
                 print("Activate the loop mode") 
                 stepTime=0.1
                 iStep=0
                 while True:
                          d=[[0,0,0]]*nLED
                          d[iStep%nLED]=[intensity]*3
                          #ws2812.write2812(spi, d)
                          iStep=(iStep+1)%nLED
                          time.sleep(stepTime)
            if mode == 'wave':
                 print("Activate the wave mode for hyper spectral imaging")
                 tStart=time.time()
                 indices=4*numpy.array(range(nLED), dtype=numpy.uint32)*numpy.pi/nLED
                 period0=2
                 period1=2.1
                 period2=2.2
                 try:
                     while True:
                          t=tStart-time.time()
                          #t=1.1
                          f=numpy.zeros((nLED,3))
                          f[:,0]=sin(2*pi*t/period0+indices)
                          f[:,1]=sin(2*pi*t/period1+indices)
                          f[:,2]=sin(2*pi*t/period2+indices)
                          f=(intensity)*((f+1.0)/2.0)
                          fi=numpy.array(f, dtype=numpy.uint8)
                          #print fi[0]
                          #time_write2812(spi, fi)
                          #ws2812.write2812(spi, fi)
                          time.sleep(0.01)
                 except KeyboardInterrupt:
                        print("Error communicating with the RGB data control")
                       #test_off(spi, nLED)
            if mode == 'npimage':
                 print("Numpy image running the loop of the led function")
def test_off(spi, nLED):
    print("Print fail safe")
    #ws2812.write2812(spi, [0,0,0]*nLED)

class Worker1(QThread):
    ImageUpdate = pyqtSignal(QImage)
    def run(self):
        self.ThreadActive = True
        Capture = cv2.VideoCapture(int(cam_num[0]))
        while self.ThreadActive:
            ret, frame = Capture.read()
            if ret:
                Image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                
                FlippedImage = cv2.flip(Image, 1)
                ConvertToQtFormat = QImage(FlippedImage.data, FlippedImage.shape[1], FlippedImage.shape[0], QImage.Format_RGB888)
                Pic = ConvertToQtFormat.scaled(531, 441, Qt.KeepAspectRatio)
                self.ImageUpdate.emit(Pic)
    def stop(self):
        self.ThreadActive = False
        self.quit()
class Worker2(QThread):
    ImageUpdate = pyqtSignal(QImage)
    def run(self):
        self.ThreadActive = True
        Capture = cv2.VideoCapture(int(cam_num[1])) 
        #detector = cv2.QRCodeDetector()
        while self.ThreadActive:
            try:
                if Qr_listdata == []:
                     Qr_listdata.append("No Qr_code_data")
                if len(Qr_listdata) > 1:
                     Qr_listdata.remove(Qr_listdata[0]) # remove the list if inside the list data has more than 1
                     print("Qr_code_Data: ",Qr_listdata[len(Qr_listdata)-1]) # Getting the data from the list here to processing data of patient loop processing 
                     #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
                     # Processing the patient data here query the data from the tube index code 
                     print(Qr_listdata[0],type(Qr_listdata[0]),Qr_listdata[0].split(",")[1])
                     Tube_index_number = Qr_listdata[0].split(",")[1]  # Extracted the data from the tube code index output 
                     print(dict_complete_query.get(int(Tube_index_number)))
                     listPatients_code = dict_complete_query.get(int(Tube_index_number))
                     tb_index = len(dict_complete_query.get(int(Tube_index_number)))
                     print(dict_complete_query)
                     for code in range(1,tb_index):
                                print(tb_index,listPatients_code[code],dict_patients_query.get(listPatients_code[code])) #Extracted patients name

                                """
                                if Tube_index_number not in list(Patient_qr_member):
                                             patient_mem_list.append([Ref_group]) # Getting the patient mem list data 
                                
                                if Patient_qr_member !={}:
                                     
                                    if Tube_index_number != list(Patient_qr_member)[len(list(Patient_qr_member))-1]:
                                                    Ref_group.clear()

                                if dict_patients_query.get(listPatients_code[code]) not in  Ref_group: 
                                                     Ref_group.append(dict_patients_query.get(listPatients_code[code]))
                                                     Groupping_Patient[0].append(Ref_group)
                                                     Patient_qr_member[Tube_index_number] = Groupping_Patient[0]
                                
                                if dict_patients_query.get(listPatients_code[code]) in Ref_group:  
                                               print("The Patient data all contained in ",Tube_index_number)
                                               
                                print(Ref_group)
                                print(list(Patient_qr_member))
                                print(patient_mem_list,len(patient_mem_list))             
                                #print(Patient_qr_member,len(list(Patient_qr_member)),len(Groupping_Patient[0])) 
                                """
                     #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
                     
            except:
                print("Out of range")     
            ret, frame = Capture.read()
            if ret:
                Image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                #print(Image)
                try:
                   
                   barcodes = pyzbar.decode(Image)
                   for barcode in barcodes:
	                              # extract the bounding box location of the barcode and draw the
	                              # bounding box surrounding the barcode on the image
	                              (x, y, w, h) = barcode.rect
                                  
	                              cv2.rectangle(Image, (x, y), (x + w, y + h), (0, 0, 255), 2)
                                  # the barcode data is a bytes object so if we want to draw it on
	                              # our output image we need to convert it to a string first
	                              barcodeData = barcode.data.decode("utf-8")
	                              barcodeType = barcode.type
	                              # draw the barcode data and barcode type on the image
	                              text = "{} ({})".format(barcodeData, barcodeType)
	                              cv2.putText(Image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX,0.5, (0, 0, 255), 2)
                                  # Add the function to output the data from hear 
	                              Qr_listdata.append("[{},{},{},{}]".format(barcodeType, barcodeData,str(len(barcodes)),(x,y,w,h))) #Grab output data 
                                  
                except:

                    print("No QRcode detected!")
                FlippedImage = cv2.flip(Image, 1)
                ConvertToQtFormat = QImage(FlippedImage.data, FlippedImage.shape[1], FlippedImage.shape[0], QImage.Format_RGB888)
                Pic = ConvertToQtFormat.scaled(541, 431, Qt.KeepAspectRatio)
                self.ImageUpdate.emit(Pic)
    def stop(self):
        self.ThreadActive = False
        self.quit()
     
def main():
    app = QtWidgets.QApplication(sys.argv)
    main = MainWindow()
    main.show()
    sys.exit(app.exec_())

if __name__ == '__main__':         
      main()
